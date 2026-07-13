#!/usr/bin/env python3
"""
从化工产品价格 Excel 表中提取最新价格，映射到公司产品。

数据源：
1. CJ (长江化工) - cj-chemical-prices-YYYYMMDD.xlsx
   - Sheet 3 (价格统计表): 包含行业、产品、单位、最新价格
2. XYZQ (兴业证券) - xyzq-chemical-prices-YYYYMMDD.xlsx
   - Sheet 0 (汇总表): 包含产品、单位、最新日期、最新价格等
   - Sheet 8-32 (各品类详细数据): 时间序列，最后一行为最新数据

输出：
- company-latest-prices.json: 公司 → 产品 → 最新价格映射
"""

import json
import os
import glob
import re
import openpyxl
from datetime import datetime


# ============================================================
# 配置
# ============================================================

SOURCES_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "terminal", "docs", "data", "sources")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "public", "data")
COMPANIES_FILE = os.path.join(OUTPUT_DIR, "companies.json")


def find_latest_file(pattern):
    """找到匹配模式的最新文件（按文件名中的日期排序）"""
    files = glob.glob(os.path.join(SOURCES_DIR, pattern))
    if not files:
        return None
    # 按文件名排序（日期在文件名末尾）
    files.sort(reverse=True)
    return files[0]


# ============================================================
# CJ 数据提取
# ============================================================

def extract_cj_prices(filepath):
    """从 CJ Excel 的价格统计表中提取最新价格"""
    print(f"\n=== 读取 CJ 数据: {os.path.basename(filepath)} ===")
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Sheet 3 是价格统计表
    ws = wb.worksheets[3]
    prices = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        industry = row[0] if row[0] else None
        product = row[1] if row[1] else None
        unit = row[2] if row[2] else None
        latest_price = row[3] if row[3] else None

        if product and latest_price:
            try:
                price_val = float(latest_price)
                prices[product] = {
                    "price": price_val,
                    "unit": str(unit) if unit else "",
                    "industry": str(industry) if industry else "",
                    "source": "cj"
                }
            except (ValueError, TypeError):
                pass

    wb.close()
    print(f"  提取到 {len(prices)} 个产品价格")
    return prices


def extract_cj_timeseries(filepath):
    """从 CJ Excel 的价格走势图(Sheet 2)中提取最新价格"""
    print(f"\n=== 读取 CJ 时间序列: {os.path.basename(filepath)} ===")
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Sheet 2 是价格走势图（时间序列）
    ws = wb.worksheets[2]
    prices = {}

    # 第1行是产品名，第2行是单位，后续行是数据
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    products = [str(h) if h else f"col_{i}" for i, h in enumerate(header_row)]

    # 找到最后一行有数据的行
    last_row = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # 第一列是日期
            last_row = row

    if last_row:
        for i, val in enumerate(last_row):
            if i > 0 and val is not None and products[i] != f"col_{i}":
                try:
                    price_val = float(val)
                    if price_val > 0:
                        prices[products[i]] = {
                            "price": price_val,
                            "source": "cj_ts"
                        }
                except (ValueError, TypeError):
                    pass

    wb.close()
    print(f"  提取到 {len(prices)} 个产品价格（时间序列）")
    return prices


# ============================================================
# XYZQ 数据提取
# ============================================================

def extract_xyzq_summary(filepath):
    """从 XYZQ Excel 的汇总表(Sheet 0)中提取最新价格"""
    print(f"\n=== 读取 XYZQ 汇总表: {os.path.basename(filepath)} ===")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    prices = {}

    # 第3行是表头: 行业, 产品, 单位, 最新日期, 最新, 周度, 月度, ...
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        industry = row[0] if row[0] else None
        product = row[1] if row[1] else None
        unit = row[2] if row[2] else None
        latest_date = row[3] if row[3] else None
        latest_price = row[4] if row[4] else None

        if product and latest_price:
            try:
                price_val = float(latest_price)
                prices[product] = {
                    "price": price_val,
                    "unit": str(unit) if unit else "",
                    "industry": str(industry) if industry else "",
                    "date": str(latest_date) if latest_date else "",
                    "source": "xyzq"
                }
            except (ValueError, TypeError):
                pass

    wb.close()
    print(f"  提取到 {len(prices)} 个产品价格")
    return prices


def extract_xyzq_detail(filepath):
    """从 XYZQ Excel 的各品类详情表(Sheet 8+)中提取最新价格"""
    print(f"\n=== 读取 XYZQ 详情表: {os.path.basename(filepath)} ===")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    prices = {}

    for sheet_idx in range(8, len(wb.sheetnames)):
        ws = wb.worksheets[sheet_idx]

        # 第1行是产品名
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        products = []
        for i, h in enumerate(header_row):
            if h and i > 1:  # 跳过前两列（图表、日期）
                products.append((i, str(h)))

        # 找最后一行有数据的行
        last_row = None
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[1] is not None:  # 第二列是日期
                last_row = row

        if last_row:
            for col_idx, product_name in products:
                val = last_row[col_idx] if col_idx < len(last_row) else None
                if val is not None:
                    try:
                        price_val = float(val)
                        if price_val > 0 and product_name not in prices:
                            prices[product_name] = {
                                "price": price_val,
                                "source": "xyzq_detail"
                            }
                    except (ValueError, TypeError):
                        pass

    wb.close()
    print(f"  提取到 {len(prices)} 个产品价格（详情表）")
    return prices


# ============================================================
# 产品名称匹配
# ============================================================

# 公司产品 commodityKey 到 Excel 产品名的映射
# 格式: { "commodityKey": ["cj产品名", "xyzq产品名"] }
KEY_TO_EXCEL_NAMES = {
    # 能源
    "cj:price:92#汽油": ["92#汽油", "中国92#汽油(镇海炼化)"],
    "cj:price:0#柴油": ["0#柴油", "中国0#柴油(镇海炼化)"],
    "cj:price:液化气": ["液化气(华东)"],
    "cj:price:LNG": ["LNG"],
    "cj:price:动力煤": ["动力煤(Q5500,秦皇岛港)"],
    "cj:price:布伦特原油": ["Brent期货(活跃合约)", "布伦特原油"],
    "cj:price:WTI原油": ["WTI期货(活跃合约)", "WTI原油"],
    "lme:price:WTI原油": ["WTI期货(活跃合约)", "WTI原油"],

    # 基础化工 - 纯碱/烧碱/氯碱
    "cj:price:轻质纯碱（华东）": ["纯碱(轻质,华东)"],
    "cj:price:重质纯碱（华东）": ["纯碱(重质,华东)"],
    "cj:price:32%隔膜烧碱（华东）": ["烧碱(32%离子膜,华北)", "烧碱(32%隔膜,华东)"],
    "cj:price:PVC（华东电石法）": ["PVC(电石法,长三角)", "PVC(华东电石法)"],
    "cj:price:电石（华东）": ["电石(西北)", "电石(华东)"],
    "cj:price:液氯（华东）": ["液氯(山东)", "液氯(华东)"],

    # 甲醇/烯烃
    "cj:price:甲醇（长三角）": ["甲醇(华东)", "甲醇(长三角)"],
    "cj:price:PP（余姚市场J340/扬子）": ["PP粒(1100,均聚注塑)", "PP(余姚市场J340/扬子)"],
    "cj:price:丙烯（汇丰石化）": ["丙烯(华东)", "丙烯(汇丰石化)"],
    "cj:price:丙烯酸": ["丙烯酸(华东)"],
    "cj:price:丙烯酸丁酯": ["丙烯酸丁酯(华东)"],
    "cj:price:环氧丙烷（华东）": ["环氧丙烷(华东)"],
    "cj:price:MTBE": ["MTBE(石大胜华)"],

    # 聚酯/涤纶
    "cj:price:PTA（华东）": ["PTA(华东)"],
    "cj:price:涤纶POY（华东）": ["涤纶长丝(POY)", "涤纶POY(华东)"],
    "cj:price:对二甲苯（PX）": ["PX(华东)"],
    "cj:price:PET切片（华东）": ["聚酯切片(纤维级,华东)", "PET切片(华东)"],
    "cj:price:乙二醇": ["乙二醇(华东)"],

    # 煤化工
    "cj:price:醋酸（华东）": ["醋酸(华东)"],
    "cj:price:DMF（华东）": ["DMF(华东)"],
    "cj:price:己二酸（华东）": ["己二酸(华东)"],
    "cj:price:丁酮（华东）": ["丁酮(华东)", "甲乙酮(华东)"],
    "cj:price:顺酐": ["顺酐(华东)"],
    "cj:price:丁二烯（上海石化）": ["丁二烯(华东)"],

    # MDI/TDI
    "cj:price:聚合MDI（华东）": ["聚合MDI(华东)"],
    "cj:price:纯MDI（华东）": ["纯MDI(华东)"],
    "cj:price:TDI（华东）": ["TDI(T80,华东)", "TDI(华东)"],
    "cj:price:PC": ["聚碳酸酯PC(科思创2805,华东)", "PC"],

    # 农药
    "cj:price:草甘膦": ["草甘膦(华东)"],
    "cj:price:草铵膦": ["草铵膦"],
    "cj:price:多菌灵": ["多菌灵(98%,华东)"],
    "cj:price:敌草快": ["敌草快(42%,华东)"],
    "cj:price:甲基硫菌灵": ["甲基硫菌灵(华东)"],
    "cj:price:敌草隆": ["敌草隆"],
    "cj:price:使它隆": ["氯氟吡氧乙酸(华东)"],
    "cj:price:毒莠定": ["毒莠定(华东)"],

    # 氟化工
    "cj:price:R32": ["R32(浙江)"],
    "cj:price:R22": ["R22(浙江)"],
    "cj:price:PVDF": ["PVDF(粉料，东岳集团)"],
    "cj:price:PTFE": ["聚四氟乙烯(浙江巨化)"],
    "cj:price:六氟化钨": ["六氟化钨"],

    # 钛白粉
    "cj:price:钛白粉(金红石)": ["钛白粉(金红石型,华东)"],
    "cj:price:钛白粉（锐钛型，攀钢钒钛）": ["钛白粉(锐钛型,华东)"],

    # 食品添加剂
    "cj:price:安赛蜜": ["安赛蜜(华东)"],
    "cj:price:三氯蔗糖": ["三氯蔗糖"],
    "cj:price:乙基麦芽酚": ["乙基麦芽酚(安徽)"],
    "cj:price:味精": ["味精(国内)"],
    "cj:price:I+G": ["I+G(华东)"],
    "cj:price:肌苷": ["肌苷(华东)"],

    # 染料
    "cj:price:分散染料": ["分散染料"],
    "cj:price:活性染料": ["活性染料"],
    "cj:price:间苯二胺": ["间苯二胺(华东)"],

    # 橡胶
    "cj:price:天然橡胶（上海地区）": ["天然橡胶(国产5号标胶)"],
    "cj:price:全钢胎": ["全钢胎"],
    "cj:price:浓缩乳胶": ["浓缩乳胶(华东)"],

    # 粘胶
    "cj:price:粘胶短纤（华东）": ["粘胶短纤(1.5D,38毫米)", "粘胶短纤(华东)"],

    # 化肥
    "cj:price:尿素(华东)": ["尿素(山东)", "尿素(华东)"],
    "cj:price:尿素（华鲁恒升小颗粒）": ["尿素(华鲁恒升小颗粒)"],
    "cj:price:复合肥": ["复合肥(45%CL,湖北)"],
    "cj:price:三聚氰胺": ["三聚氰胺(四川)"],
    "cj:price:氯化钾": ["氯化钾(60%粉,青海挂牌)"],
    "cj:price:碳酸锂": ["碳酸锂(电池级,江苏)"],

    # 其他
    "cj:price:环氧树脂": ["环氧树脂(E51,华东)"],
    "cj:price:防老剂": ["防老剂4020(华北)"],
    "cj:price:PA66": ["PA66长丝(华东)"],
    "cj:price:蛋氨酸": ["固体蛋氨酸(山东)"],
    "cj:price:VA": ["维生素A(50万IU/g,国产)"],
    "cj:price:VE": ["维生素E(50%,国产)"],
    "cj:price:苯酚(华东)": ["苯酚(华东)"],
    "cj:price:丙酮(华东)": ["丙酮(华东)"],
    "cj:price:双酚A": ["双酚A(华东)"],
    "cj:price:甲醛（长三角）": ["甲醛(华东)"],
    "cj:price:多聚甲醛": ["多聚甲醛(华东)"],
    "cj:price:硫酸": ["硫酸(98%,长三角)"],
    "cj:price:双氧水": ["双氧水(27.5%,山东)"],
    "cj:price:三氯乙烯": ["三氯乙烯(华东)"],
    "cj:price:聚乙烯醇PVA": ["PVA(1799,四川维尼纶)", "聚乙烯醇PVA"],
    "cj:price:醋酸乙烯（华东）": ["醋酸乙烯(华东)"],
    "cj:price:木屑颗粒": ["木屑颗粒"],
    "cj:price:KH-550": ["KH-550"],
    "cj:price:气相白炭黑": ["气相白炭黑"],
    "cj:price:107胶": ["107胶"],
    "cj:price:钴酸锂": ["钴酸锂"],
    "cj:price:NCM811": ["NCM811"],
    "cj:price:磷酸铁锂": ["磷酸铁锂"],
    "cj:price:磷酸铁": ["磷酸铁"],
    "cj:price:五氧化二钒": ["五氧化二钒"],
    "cj:price:钛精矿": ["钛精矿(攀枝花)"],
    "cj:price:仲钨酸铵": ["仲钨酸铵"],
    "cj:price:碳化钨": ["碳化钨"],
    "cj:price:钼精矿": ["钼精矿"],
    "cj:price:金属铟": ["铟(4N,华东)"],
    "cj:price:金属铍": ["铍"],
    "cj:price:金属镓": ["镓"],
    "cj:price:锑锭": ["锑锭(华东)"],
    "cj:price:氧化镨钕": ["氧化镨钕"],
    "cj:price:N35钕铁硼": ["N35钕铁硼"],
    "cj:price:氧化铈": ["氧化铈"],
    "cj:price:铁精粉": ["铁精粉"],
    "cj:price:铜杆": ["铜杆"],
    "cj:price:聚烯烃": ["LLDPE(余姚市场7042/扬子石化)", "聚烯烃"],
    "cj:price:豆粕": ["豆粕"],
    "cj:price:生猪": ["生猪"],
    "cj:price:葡萄糖": ["葡萄糖(国内)"],
    "cj:price:赖氨酸": ["赖氨酸(98.5%,国产)"],
    "cj:price:硬脂酸": ["硬脂酸"],
    "cj:price:磺酸": ["磺酸"],
    "cj:price:AES": ["AES"],
    "cj:price:还原剂": ["还原剂"],
    "cj:price:色酚": ["色酚"],
    "cj:price:永固紫": ["永固紫"],
    "cj:price:聚氨酯": ["聚氨酯"],
    "cj:price:薄荷醇": ["薄荷醇"],
    "cj:price:PPS": ["PPS"],
    "cj:price:三氯甲醛": ["三氯甲醛"],
    "cj:price:DMC": ["DMC(华东)", "二甲基环硅氧烷(DMC)(华东)"],
    "cj:price:石脑油（中石化）": ["石脑油(中石化)"],
    "cj:price:NYMEX天然气": ["NYMEX天然气"],
    "cj:price:氧氯化锆": ["氧氯化锆"],
    "cj:price:四氧化三锰": ["四氧化三锰"],
    "cj:price:乙醇": ["乙醇(普通,山东)"],
    "cj:price:玉米淀粉": ["玉米淀粉(滨州,出厂价)"],
    "cj:price:电子雷管": ["电子雷管"],
    "cj:price:工业炸药": ["工业炸药"],
    "cj:price:硝酸铵": ["硝酸铵(山东)"],
    "cj:price:高氯酸钾": ["高氯酸钾"],
    "cj:price:风电机组": ["风电机组"],

    # LME 金属
    "lme:price:铜": ["铜(现货)", "铜"],
    "lme:price:铝": ["铝(现货)", "铝"],
    "lme:price:锌": ["锌(现货)", "锌"],
    "lme:price:锡": ["锡(现货)", "锡"],
    "lme:price:镍": ["镍(现货)", "镍"],
    "lme:price:铅": ["铅(现货)", "铅"],
    "lme:price:黄金": ["黄金(现货)", "黄金"],
    "lme:price:白银": ["白银(现货)", "白银"],
    "lme:price:铂": ["铂(现货)", "铂"],
    "lme:price:钯": ["钯(现货)", "钯"],
    "lme:price:钴": ["钴(现货)", "钴"],
    "lme:price:铌": ["铌(现货)", "铌"],
    "lme:price:钽": ["钽(现货)", "钽"],
    "lme:price:铁矿石": ["铁矿石(现货)", "铁矿石"],
    "lme:price:氧化铝": ["氧化铝(现货)", "氧化铝"],
    "lme:price:铑": ["铑"],

    # 其他 cj
    "cj:price:氯化铵": ["氯化铵"],
    "cj:price:硫酸钾": ["硫酸钾"],
    "cj:price:二甲基环硅氧烷(DMC)(华东)": ["二甲基环硅氧烷(DMC)(华东)"],
    "cj:price:萤石粉": ["萤石粉(湿粉,华东)"],
    "cj:price:无水氢氟酸": ["无水氢氟酸(华东)"],
    "cj:price:液氨": ["液氨(山东)"],
    "cj:price:己内酰胺（CPL）": ["己内酰胺(CPL)", "己内酰胺（CPL）"],
    "cj:price:丙烯腈": ["丙烯腈"],
    "cj:price:氨纶40D（华东）": ["氨纶40D(华东)"],
    "cj:price:PVA": ["PVA(1799,四川维尼纶)"],
    "cj:price:醋酸乙烯": ["醋酸乙烯(华东)"],

    # xyzq 特有的 key
    "xyzq:price:复合肥(45%CL,湖北)": ["复合肥(45%CL,湖北)"],
    "xyzq:price:磷酸铁": ["磷酸铁"],

    # 补充映射 - 从 Excel 实际名称反查
    "cj:price:使它隆": ["氯氟吡氧乙酸(华东)", "使它隆"],
    "cj:price:毒莠定": ["毒莠定(华东)", "毒莠定"],
    "cj:price:敌草快": ["敌草快(42%,华东)", "敌草快"],
    "cj:price:安赛蜜": ["安赛蜜(华东)"],
    "cj:price:乙基麦芽酚": ["乙基麦芽酚(安徽)"],
    "cj:price:AES": ["AES(华东)"],
    "cj:price:硬脂酸": ["硬脂酸(华东)"],
    "cj:price:磺酸": ["磺酸(华东)"],
    "cj:price:双氧水": ["双氧水(27.5%,山东)"],
    "cj:price:聚氨酯": ["聚氨酯(华东)"],
    "cj:price:色酚": ["色酚(华东)"],
    "cj:price:永固紫": ["永固紫(华东)"],
    "cj:price:PTFE": ["聚四氟乙烯(浙江巨化)", "PTFE"],
    "cj:price:PVDF": ["PVDF(粉料，东岳集团)", "PVDF(涂料用,东岳集团)", "PVDF(电池用,东岳集团)"],
    "cj:price:PA66": ["PA66长丝(华东)"],
    "cj:price:VA": ["维生素A(50万IU/g,国产)"],
    "cj:price:VE": ["维生素E(50%,国产)"],
    "cj:price:MDI": ["聚合MDI(华东)", "聚合MDI"],
    "cj:price:纯MDI（华东）": ["纯MDI(华东)"],
    "cj:price:浓缩乳胶": ["浓缩乳胶(华东)"],
    "cj:price:薄荷醇": ["薄荷醇(华东)"],
    "cj:price:PPS": ["PPS(华东)"],
    "cj:price:KH-550": ["KH-550(华东)"],
    "cj:price:107胶": ["107胶(华东)"],
    "cj:price:钴酸锂": ["钴酸锂(华东)"],
    "cj:price:NCM811": ["NCM811(华东)"],
    "cj:price:磷酸铁锂": ["磷酸铁锂(华东)", "磷酸铁锂电解液"],
    "cj:price:电子雷管": ["电子雷管(华东)"],
    "cj:price:工业炸药": ["工业炸药(华东)"],
    "cj:price:硝酸铵": ["硝酸铵(山东)"],
    "cj:price:高氯酸钾": ["高氯酸钾(华东)"],
    "cj:price:风电机组": ["风电机组(华东)"],
    "cj:price:木屑颗粒": ["木屑颗粒(华东)"],
    "cj:price:还原剂": ["还原剂(华东)"],
    "cj:price:I+G": ["I+G(华东)"],
    "cj:price:肌苷": ["肌苷(华东)"],
    "cj:price:靛蓝": ["靛蓝(华东)"],
    "cj:price:间苯二胺": ["间苯二胺(华东)"],
    "cj:price:甲基麦芽酚": ["甲基麦芽酚(安徽)"],

    # LME 金属补充
    "lme:price:铜": ["铜(LME)", "铜(现货)", "铜"],
    "lme:price:铝": ["铝(LME)", "铝(现货)", "铝"],
    "lme:price:锌": ["锌(LME)", "锌(现货)", "锌"],
    "lme:price:锡": ["锡(LME)", "锡(现货)", "锡"],
    "lme:price:镍": ["镍(LME)", "镍(现货)", "镍"],
    "lme:price:铅": ["铅(LME)", "铅(现货)", "铅"],
    "lme:price:黄金": ["黄金(LME)", "黄金(现货)", "黄金"],
    "lme:price:白银": ["白银(LME)", "白银(现货)", "白银"],
    "lme:price:铂": ["铂(LME)", "铂(现货)", "铂"],
    "lme:price:钯": ["钯(LME)", "钯(现货)", "钯"],
    "lme:price:钴": ["钴(LME)", "钴(现货)", "钴"],
    "lme:price:铌": ["铌(LME)", "铌(现货)", "铌"],
    "lme:price:钽": ["钽(LME)", "钽(现货)", "钽"],
    "lme:price:铁矿石": ["铁矿石(LME)", "铁矿石(现货)", "铁矿石"],
    "lme:price:氧化铝": ["氧化铝(LME)", "氧化铝(现货)", "氧化铝"],
    "lme:price:铑": ["铑(LME)", "铑(现货)", "铑"],

    # 其他补充
    "cj:price:P.O42.5水泥": ["P.O42.5水泥"],
    "cj:price:氧氯化锆": ["氧氯化锆(华东)"],
    "cj:price:四氧化三锰": ["四氧化三锰(华东)"],
    "cj:price:仲钨酸铵": ["仲钨酸铵(华东)"],
    "cj:price:碳化钨": ["碳化钨(华东)"],
    "cj:price:钼精矿": ["钼精矿(华东)"],
    "cj:price:金属铟": ["铟(4N,华东)"],
    "cj:price:金属铍": ["铍(华东)"],
    "cj:price:金属镓": ["镓(华东)"],
    "cj:price:锑锭": ["锑锭(华东)"],
    "cj:price:氧化镨钕": ["氧化镨钕(华东)"],
    "cj:price:N35钕铁硼": ["N35钕铁硼(华东)"],
    "cj:price:氧化铈": ["氧化铈(华东)"],
    "cj:price:铁精粉": ["铁精粉(华东)"],
    "cj:price:铜杆": ["铜杆(华东)"],
    "cj:price:硫酸": ["硫酸(98%,长三角)"],
    "cj:price:三氯乙烯": ["三氯乙烯(华东)"],
    "cj:price:聚乙烯醇PVA": ["PVA(1799,四川维尼纶)", "聚乙烯醇PVA"],
    "cj:price:醋酸乙烯（华东）": ["醋酸乙烯(华东)"],
    "cj:price:己内酰胺（CPL）": ["己内酰胺(CPL)"],
    "cj:price:丙烯腈": ["丙烯腈(华东)"],
    "cj:price:氨纶40D（华东）": ["氨纶40D(华东)"],
    "cj:price:石脑油（中石化）": ["石脑油(中石化)"],
    "cj:price:DMC": ["DMC(华东)", "二甲基环硅氧烷(DMC)(华东)"],
    "cj:price:氯化铵": ["氯化铵(农湿,江苏华昌)"],
    "cj:price:硫酸钾": ["硫酸钾(50%,山东)"],
    "cj:price:萤石粉": ["萤石粉(湿粉,华东)"],
    "cj:price:无水氢氟酸": ["无水氢氟酸(华东)"],
    "cj:price:液氨": ["液氨(山东)"],
    "cj:price:乙醇": ["乙醇(普通,山东)"],
    "cj:price:玉米淀粉": ["玉米淀粉(滨州,出厂价)"],
}


def find_price_for_key(commodity_key, cj_prices, xyzq_prices, cj_ts_prices, xyzq_detail_prices):
    """根据 commodityKey 查找最新价格"""
    # 先查映射表
    excel_names = KEY_TO_EXCEL_NAMES.get(commodity_key, [])

    # 也用 key 本身提取产品名
    key_product = commodity_key.split(":")[-1] if ":" in commodity_key else commodity_key

    # 搜索顺序: xyzq_detail > xyzq > cj_ts > cj
    all_sources = [
        ("xyzq_detail", xyzq_detail_prices),
        ("xyzq", xyzq_prices),
        ("cj_ts", cj_ts_prices),
        ("cj", cj_prices),
    ]

    # 先用映射名搜索
    for name in excel_names:
        for source_name, source_prices in all_sources:
            if name in source_prices:
                return {
                    "price": source_prices[name]["price"],
                    "source": source_name,
                    "matched_name": name,
                    "unit": source_prices[name].get("unit", ""),
                }

    # 再用 key 产品名搜索
    for source_name, source_prices in all_sources:
        if key_product in source_prices:
            return {
                "price": source_prices[key_product]["price"],
                "source": source_name,
                "matched_name": key_product,
                "unit": source_prices[key_product].get("unit", ""),
            }

    # 模糊匹配 - 用 key 产品名在所有价格中搜索
    for source_name, source_prices in all_sources:
        for sp_name, sp_data in source_prices.items():
            # 双向包含匹配
            if key_product in sp_name or sp_name in key_product:
                return {
                    "price": sp_data["price"],
                    "source": source_name,
                    "matched_name": sp_name,
                    "unit": sp_data.get("unit", ""),
                }

    # 更宽松的模糊匹配 - 去除括号内容后比较
    clean_key = re.sub(r'[（(][^）)]*[）)]', '', key_product).strip()
    if clean_key and len(clean_key) >= 2:
        for source_name, source_prices in all_sources:
            for sp_name, sp_data in source_prices.items():
                clean_sp = re.sub(r'[（(][^）)]*[）)]', '', sp_name).strip()
                if clean_key in clean_sp or clean_sp in clean_key:
                    return {
                        "price": sp_data["price"],
                        "source": source_name,
                        "matched_name": sp_name,
                        "unit": sp_data.get("unit", ""),
                    }

    return None


# ============================================================
# 主流程
# ============================================================

def main():
    print("=" * 60)
    print("化工产品最新价格提取工具")
    print("=" * 60)

    # 1. 找到最新的 Excel 文件
    cj_file = find_latest_file("cj-chemical-prices-*.xlsx")
    xyzq_file = find_latest_file("xyzq-chemical-prices-*.xlsx")

    if not cj_file:
        print("错误: 未找到 CJ 价格文件")
        return
    if not xyzq_file:
        print("错误: 未找到 XYZQ 价格文件")
        return

    print(f"\nCJ 文件: {os.path.basename(cj_file)}")
    print(f"XYZQ 文件: {os.path.basename(xyzq_file)}")

    # 2. 提取价格
    cj_prices = extract_cj_prices(cj_file)
    cj_ts_prices = extract_cj_timeseries(cj_file)
    xyzq_prices = extract_xyzq_summary(xyzq_file)
    xyzq_detail_prices = extract_xyzq_detail(xyzq_file)

    # 合并所有价格到一个查找表
    all_prices = {}
    for source in [cj_prices, cj_ts_prices, xyzq_prices, xyzq_detail_prices]:
        for name, data in source.items():
            if name not in all_prices:
                all_prices[name] = data

    print(f"\n总计提取到 {len(all_prices)} 个不同产品价格")

    # 3. 读取公司映射
    with open(COMPANIES_FILE, "r", encoding="utf-8") as f:
        companies = json.load(f)

    # 4. 为每个公司产品查找最新价格
    results = []
    total_products = 0
    matched_products = 0

    for company in companies:
        code = company["code"]
        name = company["name"]
        group = company["group"]

        company_result = {
            "code": code,
            "name": name,
            "group": group,
            "products": []
        }

        for product in company.get("productsByRevenue", []):
            total_products += 1
            commodity_key = product["commodityKey"]
            product_name = product["productName"]

            if commodity_key.startswith("unknown:"):
                company_result["products"].append({
                    "productName": product_name,
                    "commodityKey": commodity_key,
                    "latestPrice": None,
                    "status": "unknown_key"
                })
                continue

            price_info = find_price_for_key(
                commodity_key, cj_prices, xyzq_prices, cj_ts_prices, xyzq_detail_prices
            )

            if price_info:
                matched_products += 1
                company_result["products"].append({
                    "productName": product_name,
                    "commodityKey": commodity_key,
                    "latestPrice": price_info["price"],
                    "unit": price_info["unit"],
                    "source": price_info["source"],
                    "matchedName": price_info["matched_name"],
                    "status": "matched"
                })
            else:
                company_result["products"].append({
                    "productName": product_name,
                    "commodityKey": commodity_key,
                    "latestPrice": None,
                    "status": "not_found"
                })

        results.append(company_result)

    # 5. 输出统计
    print(f"\n{'=' * 60}")
    print(f"映射结果统计")
    print(f"{'=' * 60}")
    print(f"公司总数: {len(companies)}")
    print(f"产品总数: {total_products}")
    print(f"成功匹配: {matched_products}")
    print(f"未匹配: {total_products - matched_products}")
    print(f"匹配率: {matched_products / total_products * 100:.1f}%")

    # 列出未匹配的产品
    not_found = []
    for company in results:
        for product in company["products"]:
            if product["status"] == "not_found":
                not_found.append(f"  {company['name']} ({company['code']}): {product['productName']} -> {product['commodityKey']}")

    if not_found:
        print(f"\n未匹配的产品 ({len(not_found)}):")
        for item in not_found:
            print(item)

    # 6. 保存结果
    output_file = os.path.join(OUTPUT_DIR, "company-latest-prices.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump({
            "metadata": {
                "generatedAt": datetime.now().isoformat(),
                "cjSource": os.path.basename(cj_file),
                "xyzqSource": os.path.basename(xyzq_file),
                "totalCompanies": len(companies),
                "totalProducts": total_products,
                "matchedProducts": matched_products,
                "matchRate": round(matched_products / total_products * 100, 1)
            },
            "companies": results
        }, f, ensure_ascii=False, indent=2)

    print(f"\n结果已保存到: {output_file}")

    # 7. 同时输出一个简洁的价格表
    price_table = {}
    for company in results:
        for product in company["products"]:
            if product["latestPrice"] is not None:
                key = product["commodityKey"]
                if key not in price_table:
                    price_table[key] = {
                        "productName": product["productName"],
                        "price": product["latestPrice"],
                        "unit": product.get("unit", ""),
                        "source": product["source"]
                    }

    price_table_file = os.path.join(OUTPUT_DIR, "price-table.json")
    with open(price_table_file, "w", encoding="utf-8") as f:
        json.dump(price_table, f, ensure_ascii=False, indent=2)

    print(f"价格表已保存到: {price_table_file}")


if __name__ == "__main__":
    main()
