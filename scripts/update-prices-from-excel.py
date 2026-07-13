#!/usr/bin/env python3
"""
化工产品价格更新脚本

使用方法：
1. 将新的 Excel 文件放到 terminal/docs/data/sources/ 目录
2. 运行此脚本: python3 scripts/update-prices-from-excel.py
3. 脚本会自动找到最新的 Excel 文件，提取价格，更新数据

支持的 Excel 文件格式：
- CJ: cj-chemical-prices-YYYYMMDD.xlsx (长江化工)
- XYZQ: xyzq-chemical-prices-YYYYMMDD.xlsx (兴业证券)

输出文件：
- public/data/company-latest-prices.json - 公司最新价格映射
- public/data/price-table.json - 价格速查表
- public/data/commodity-prices.json - 商品价格时间序列（增量更新）
"""

import json
import os
import sys
import glob
import re
from datetime import datetime

# 添加 openpyxl 依赖检查
try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 路径配置
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
SOURCES_DIR = os.path.join(PROJECT_DIR, "..", "terminal", "docs", "data", "sources")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "public", "data")
COMPANIES_FILE = os.path.join(OUTPUT_DIR, "companies.json")


def find_latest_file(pattern):
    """找到匹配模式的最新文件"""
    files = glob.glob(os.path.join(SOURCES_DIR, pattern))
    if not files:
        return None
    files.sort(reverse=True)
    return files[0]


# ============================================================
# Excel 数据提取
# ============================================================

def extract_cj_latest(filepath):
    """从 CJ Excel 提取最新价格"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    prices = {}

    # 方法1: 价格统计表 (Sheet 3)
    ws = wb.worksheets[3]
    for row in ws.iter_rows(min_row=2, values_only=True):
        product = row[1]
        unit = row[2]
        latest_price = row[3]
        if product and latest_price:
            try:
                prices[str(product)] = {
                    "price": float(latest_price),
                    "unit": str(unit) if unit else "",
                    "source": "cj"
                }
            except (ValueError, TypeError):
                pass

    # 方法2: 时间序列最后一行 (Sheet 2)
    ws2 = wb.worksheets[2]
    header = list(ws2.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    product_names = [str(h) if h else f"col_{i}" for i, h in enumerate(header)]

    last_row = None
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            last_row = row

    if last_row:
        for i, val in enumerate(last_row):
            if i > 0 and val is not None and product_names[i] != f"col_{i}":
                try:
                    price_val = float(val)
                    if price_val > 0 and product_names[i] not in prices:
                        prices[product_names[i]] = {
                            "price": price_val,
                            "source": "cj_ts"
                        }
                except (ValueError, TypeError):
                    pass

    wb.close()
    return prices


def extract_xyzq_latest(filepath):
    """从 XYZQ Excel 提取最新价格"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    prices = {}

    # 方法1: 汇总表 (Sheet 0)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=4, values_only=True):
        product = row[1]
        unit = row[2]
        latest_price = row[4]  # 第5列是最新价格
        if product and latest_price:
            try:
                prices[str(product)] = {
                    "price": float(latest_price),
                    "unit": str(unit) if unit else "",
                    "source": "xyzq"
                }
            except (ValueError, TypeError):
                pass

    # 方法2: 各品类详情表 (Sheet 8+)
    for sheet_idx in range(8, len(wb.sheetnames)):
        ws = wb.worksheets[sheet_idx]
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        products = []
        for i, h in enumerate(header):
            if h and i > 1:
                products.append((i, str(h)))

        last_row = None
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[1] is not None:
                last_row = row

        if last_row:
            for col_idx, name in products:
                val = last_row[col_idx] if col_idx < len(last_row) else None
                if val is not None:
                    try:
                        price_val = float(val)
                        if price_val > 0 and name not in prices:
                            prices[name] = {
                                "price": price_val,
                                "source": "xyzq_detail"
                            }
                    except (ValueError, TypeError):
                        pass

    wb.close()
    return prices


# ============================================================
# 产品名称匹配
# ============================================================

# 完整的 commodityKey -> Excel 产品名映射
KEY_TO_EXCEL = {
    # 能源
    "cj:price:92#汽油": ["92#汽油", "中国92#汽油(镇海炼化)"],
    "cj:price:0#柴油": ["0#柴油", "中国0#柴油(镇海炼化)"],
    "cj:price:液化气": ["液化气(华东)"],
    "cj:price:LNG": ["LNG"],
    "cj:price:动力煤": ["动力煤(Q5500,秦皇岛港)"],
    "cj:price:布伦特原油": ["Brent期货(活跃合约)", "布伦特原油"],
    "cj:price:WTI原油": ["WTI期货(活跃合约)", "WTI原油"],
    "cj:price:石脑油（中石化）": ["石脑油(中石化)"],
    "cj:price:NYMEX天然气": ["NYMEX天然气"],
    "lme:price:WTI原油": ["WTI期货(活跃合约)", "WTI原油"],

    # 纯碱/烧碱/氯碱
    "cj:price:轻质纯碱（华东）": ["纯碱(轻质,华东)"],
    "cj:price:重质纯碱（华东）": ["纯碱(重质,华东)"],
    "cj:price:32%隔膜烧碱（华东）": ["烧碱(32%离子膜,华北)", "烧碱(32%隔膜,华东)", "隔膜烧碱(华东)"],
    "cj:price:PVC（华东电石法）": ["PVC(电石法,长三角)", "PVC(华东电石法)", "电石法PVC"],
    "cj:price:电石（华东）": ["电石(西北)", "电石(华东)"],
    "cj:price:液氯（华东）": ["液氯(山东)", "液氯(华东)"],
    "cj:price:硫酸": ["硫酸(98%,长三角)"],
    "cj:price:双氧水": ["双氧水(27.5%,山东)"],
    "cj:price:三氯乙烯": ["三氯乙烯(华东)"],

    # 甲醇/烯烃
    "cj:price:甲醇（长三角）": ["甲醇(华东)", "甲醇(长三角)"],
    "cj:price:PP（余姚市场J340/扬子）": ["PP粒(1100,均聚注塑)", "PP(余姚市场J340/扬子)"],
    "cj:price:丙烯（汇丰石化）": ["丙烯(华东)", "丙烯(汇丰石化)"],
    "cj:price:丙烯酸": ["丙烯酸(华东)"],
    "cj:price:丙烯酸丁酯": ["丙烯酸丁酯(华东)"],
    "cj:price:环氧丙烷（华东）": ["环氧丙烷(华东)"],
    "cj:price:MTBE": ["MTBE(石大胜华)"],
    "cj:price:丙烯腈": ["丙烯腈(华东)"],

    # 聚酯/涤纶
    "cj:price:PTA（华东）": ["PTA(华东)"],
    "cj:price:涤纶POY（华东）": ["涤纶长丝(POY)", "涤纶POY(华东)"],
    "cj:price:对二甲苯（PX）": ["PX(华东)"],
    "cj:price:PET切片（华东）": ["聚酯切片(纤维级,华东)", "PET切片(华东)"],
    "cj:price:乙二醇": ["乙二醇(华东)"],

    # 煤化工
    "cj:price:醋酸（华东）": ["醋酸(华东)"],
    "cj:price:DMF（华东）": ["DMF(华东)"],
    "cj:price:己二酸（华东）": ["己二酸(华东)"],
    "cj:price:丁酮（华东）": ["丁酮(华东)", "甲乙酮(华东)"],
    "cj:price:顺酐": ["顺酐(华东)"],
    "cj:price:丁二烯（上海石化）": ["丁二烯(华东)"],
    "cj:price:BDO": ["BDO(华东)"],
    "cj:price:己内酰胺（CPL）": ["己内酰胺(CPL)"],

    # MDI/TDI
    "cj:price:聚合MDI（华东）": ["聚合MDI(华东)", "聚合MDI"],
    "cj:price:纯MDI（华东）": ["纯MDI(华东)"],
    "cj:price:TDI（华东）": ["TDI(T80,华东)", "TDI(华东)"],
    "cj:price:PC": ["聚碳酸酯PC(科思创2805,华东)", "PC"],

    # 农药
    "cj:price:草甘膦": ["草甘膦(华东)"],
    "cj:price:草铵膦": ["草铵膦"],
    "cj:price:多菌灵": ["多菌灵(98%,华东)"],
    "cj:price:敌草快": ["敌草快(42%,华东)"],
    "cj:price:甲基硫菌灵": ["甲基硫菌灵(华东)"],
    "cj:price:敌草隆": ["敌草隆"],
    "cj:price:使它隆": ["氯氟吡氧乙酸(华东)", "使它隆"],
    "cj:price:毒莠定": ["毒莠定(华东)", "毒莠定"],
    "cj:price:间苯二胺": ["间苯二胺(华东)"],

    # 氟化工
    "cj:price:R32": ["R32(浙江)"],
    "cj:price:R22": ["R22(浙江)"],
    "cj:price:PVDF": ["PVDF(粉料，东岳集团)", "PVDF(涂料用,东岳集团)"],
    "cj:price:PTFE": ["聚四氟乙烯(浙江巨化)", "PTFE"],
    "cj:price:六氟化钨": ["六氟化钨"],
    "cj:price:萤石粉": ["萤石粉(湿粉,华东)"],
    "cj:price:无水氢氟酸": ["无水氢氟酸(华东)"],

    # 钛白粉
    "cj:price:钛白粉(金红石)": ["钛白粉(金红石型,华东)"],
    "cj:price:钛白粉（锐钛型，攀钢钒钛）": ["钛白粉(锐钛型,华东)"],

    # 食品添加剂
    "cj:price:安赛蜜": ["安赛蜜(华东)"],
    "cj:price:三氯蔗糖": ["三氯蔗糖"],
    "cj:price:乙基麦芽酚": ["乙基麦芽酚(安徽)"],
    "cj:price:味精": ["味精(国内)"],
    "cj:price:I+G": ["I+G(华东)"],
    "cj:price:肌苷": ["肌苷(华东)"],
    "cj:price:蛋氨酸": ["固体蛋氨酸(山东)"],
    "cj:price:VA": ["维生素A(50万IU/g,国产)"],
    "cj:price:VE": ["维生素E(50%,国产)"],

    # 染料
    "cj:price:分散染料": ["分散染料"],
    "cj:price:活性染料": ["活性染料"],
    "cj:price:还原剂": ["还原剂"],
    "cj:price:靛蓝": ["靛蓝"],
    "cj:price:色酚": ["色酚"],
    "cj:price:永固紫": ["永固紫"],

    # 橡胶
    "cj:price:天然橡胶（上海地区）": ["天然橡胶(国产5号标胶)"],
    "cj:price:全钢胎": ["全钢胎"],
    "cj:price:浓缩乳胶": ["浓缩乳胶(华东)"],

    # 粘胶
    "cj:price:粘胶短纤（华东）": ["粘胶短纤(1.5D,38毫米)", "粘胶短纤(华东)"],

    # 化肥
    "cj:price:尿素(华东)": ["尿素(山东)", "尿素(华东)"],
    "cj:price:尿素（华鲁恒升小颗粒）": ["尿素(华鲁恒升小颗粒)"],
    "cj:price:复合肥": ["复合肥(45%CL,湖北)"],
    "cj:price:三聚氰胺": ["三聚氰胺(四川)"],
    "cj:price:氯化钾": ["氯化钾(60%粉,青海挂牌)"],
    "cj:price:碳酸锂": ["碳酸锂(电池级,江苏)"],
    "cj:price:氯化铵": ["氯化铵(农湿,江苏华昌)"],
    "cj:price:硫酸钾": ["硫酸钾(50%,山东)"],
    "cj:price:液氨": ["液氨(山东)"],

    # 树脂/助剂
    "cj:price:环氧树脂": ["环氧树脂(E51,华东)"],
    "cj:price:防老剂": ["防老剂4020(华北)"],
    "cj:price:PA66": ["PA66长丝(华东)"],
    "cj:price:聚乙烯醇PVA": ["PVA(1799,四川维尼纶)", "聚乙烯醇PVA"],
    "cj:price:醋酸乙烯（华东）": ["醋酸乙烯(华东)"],
    "cj:price:聚氨酯": ["聚氨酯"],
    "cj:price:聚烯烃": ["LLDPE(余姚市场7042/扬子石化)"],

    # 其他化工品
    "cj:price:苯酚(华东)": ["苯酚(华东)"],
    "cj:price:丙酮(华东)": ["丙酮(华东)"],
    "cj:price:双酚A": ["双酚A(华东)"],
    "cj:price:甲醛（长三角）": ["甲醛(华东)"],
    "cj:price:多聚甲醛": ["多聚甲醛(华东)"],
    "cj:price:DMC": ["DMC(华东)", "二甲基环硅氧烷(DMC)(华东)", "碳酸二甲酯DMC(华东)"],
    "cj:price:KH-550": ["KH-550"],
    "cj:price:气相白炭黑": ["气相白炭黑"],
    "cj:price:107胶": ["107胶"],
    "cj:price:薄荷醇": ["薄荷醇"],
    "cj:price:PPS": ["PPS"],
    "cj:price:三氯甲醛": ["三氯甲醛"],
    "cj:price:氨纶40D（华东）": ["氨纶40D(华东)", "氨纶(40D)"],

    # 锂电材料
    "cj:price:钴酸锂": ["钴酸锂"],
    "cj:price:NCM811": ["NCM811"],
    "cj:price:磷酸铁锂": ["磷酸铁锂"],
    "cj:price:磷酸铁": ["磷酸铁"],

    # 爆炸物
    "cj:price:硝酸铵": ["硝酸铵(山东)"],
    "cj:price:电子雷管": ["电子雷管"],
    "cj:price:工业炸药": ["工业炸药"],
    "cj:price:高氯酸钾": ["高氯酸钾"],

    # 能源/其他
    "cj:price:木屑颗粒": ["木屑颗粒"],
    "cj:price:风电机组": ["风电机组"],
    "cj:price:P.O42.5水泥": ["P.O42.5水泥"],
    "cj:price:豆粕": ["豆粕"],
    "cj:price:生猪": ["生猪"],
    "cj:price:葡萄糖": ["葡萄糖(国内)"],
    "cj:price:赖氨酸": ["赖氨酸(98.5%,国产)"],
    "cj:price:硬脂酸": ["硬脂酸"],
    "cj:price:磺酸": ["磺酸"],
    "cj:price:AES": ["AES"],
    "cj:price:乙醇": ["乙醇(普通,山东)"],
    "cj:price:玉米淀粉": ["玉米淀粉(滨州,出厂价)"],
    "cj:price:石脑油（中石化）": ["石脑油(中石化)"],
    "cj:price:氧氯化锆": ["氧氯化锆"],
    "cj:price:四氧化三锰": ["四氧化三锰"],

    # 稀土/金属（化工 Excel 可能有的）
    "cj:price:氧化镨钕": ["氧化镨钕"],
    "cj:price:N35钕铁硼": ["N35钕铁硼"],
    "cj:price:氧化铈": ["氧化铈"],
    "cj:price:仲钨酸铵": ["仲钨酸铵"],
    "cj:price:碳化钨": ["碳化钨"],
    "cj:price:钼精矿": ["钼精矿"],
    "cj:price:金属铟": ["铟(4N,华东)"],
    "cj:price:金属铍": ["铍"],
    "cj:price:金属镓": ["镓"],
    "cj:price:锑锭": ["锑锭(华东)"],
    "cj:price:五氧化二钒": ["五氧化二钒"],
    "cj:price:钛精矿": ["钛精矿(攀枝花)"],
    "cj:price:铁精粉": ["铁精粉"],
    "cj:price:铜杆": ["铜杆"],

    # LME 金属
    "lme:price:铜": ["铜(LME)", "铜(现货)", "铜"],
    "lme:price:铝": ["铝(LME)", "铝(现货)", "铝"],
    "lme:price:锌": ["锌(LME)", "锌(现货)", "锌"],
    "lme:price:锡": ["锡(LME)", "锡(现货)", "锡"],
    "lme:price:镍": ["镍(LME)", "镍(现货)", "镍"],
    "lme:price:铅": ["铅(LME)", "铅(现货)", "铅"],
    "lme:price:黄金": ["黄金(LME)", "黄金(现货)", "黄金"],
    "lme:price:白银": ["白银(LME)", "白银(现货)", "白银"],
    "lme:price:铂": ["铂(LME)", "铂(现货)", "铂"],
    "lme:price:钯": ["钯(LME)", "钯(现货)", "钯"],
    "lme:price:钴": ["钴(LME)", "钴(现货)", "钴"],
    "lme:price:铌": ["铌(LME)", "铌(现货)", "铌"],
    "lme:price:钽": ["钽(LME)", "钽(现货)", "钽"],
    "lme:price:铁矿石": ["铁矿石(LME)", "铁矿石(现货)", "铁矿石"],
    "lme:price:氧化铝": ["氧化铝(LME)", "氧化铝(现货)", "氧化铝"],
    "lme:price:铑": ["铑(LME)", "铑(现货)", "铑"],

    # xyzq 特有
    "xyzq:price:复合肥(45%CL,湖北)": ["复合肥(45%CL,湖北)"],
    "xyzq:price:磷酸铁": ["磷酸铁"],
}


def fuzzy_match(key_product, name):
    """模糊匹配两个产品名称"""
    # 精确匹配
    if key_product == name:
        return True
    # 包含匹配
    if key_product in name or name in key_product:
        return True
    # 去除括号后匹配
    clean_key = re.sub(r'[（(][^）)]*[）)]', '', key_product).strip()
    clean_name = re.sub(r'[（(][^）)]*[）)]', '', name).strip()
    if clean_key and clean_name and (clean_key in clean_name or clean_name in clean_key):
        return True
    return False


def find_price(commodity_key, all_prices):
    """为 commodityKey 查找最新价格"""
    # 1. 查映射表
    excel_names = KEY_TO_EXCEL.get(commodity_key, [])
    for name in excel_names:
        if name in all_prices:
            return all_prices[name]

    # 2. 用 key 产品名直接查
    key_product = commodity_key.split(":")[-1] if ":" in commodity_key else commodity_key
    if key_product in all_prices:
        return all_prices[key_product]

    # 3. 模糊匹配
    for name in excel_names:
        for sp_name, sp_data in all_prices.items():
            if fuzzy_match(name, sp_name):
                return sp_data

    # 4. 用 key 产品名模糊匹配
    for sp_name, sp_data in all_prices.items():
        if fuzzy_match(key_product, sp_name):
            return sp_data

    return None


# ============================================================
# 主流程
# ============================================================

def main():
    print("=" * 60)
    print("化工产品价格更新工具")
    print(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 1. 检查数据源目录
    if not os.path.exists(SOURCES_DIR):
        print(f"错误: 数据源目录不存在: {SOURCES_DIR}")
        print("请将 Excel 文件放到该目录")
        return

    # 2. 找到最新文件
    cj_file = find_latest_file("cj-chemical-prices-*.xlsx")
    xyzq_file = find_latest_file("xyzq-chemical-prices-*.xlsx")

    print(f"\n数据源文件:")
    print(f"  CJ: {os.path.basename(cj_file) if cj_file else '未找到'}")
    print(f"  XYZQ: {os.path.basename(xyzq_file) if xyzq_file else '未找到'}")

    if not cj_file and not xyzq_file:
        print("\n错误: 未找到任何价格 Excel 文件")
        return

    # 3. 提取价格
    all_prices = {}

    if cj_file:
        print(f"\n读取 CJ 数据...")
        cj_prices = extract_cj_latest(cj_file)
        all_prices.update(cj_prices)
        print(f"  CJ 产品数: {len(cj_prices)}")

    if xyzq_file:
        print(f"\n读取 XYZQ 数据...")
        xyzq_prices = extract_xyzq_latest(xyzq_file)
        all_prices.update(xyzq_prices)
        print(f"  XYZQ 产品数: {len(xyzq_prices)}")

    print(f"\n合并后总产品数: {len(all_prices)}")

    # 4. 读取公司映射
    if not os.path.exists(COMPANIES_FILE):
        print(f"错误: 公司映射文件不存在: {COMPANIES_FILE}")
        return

    with open(COMPANIES_FILE, "r", encoding="utf-8") as f:
        companies = json.load(f)

    # 5. 匹配价格
    results = []
    total = 0
    matched = 0
    not_found_list = []

    for company in companies:
        code = company["code"]
        name = company["name"]
        group = company["group"]

        company_data = {
            "code": code,
            "name": name,
            "group": group,
            "products": []
        }

        for product in company.get("productsByRevenue", []):
            total += 1
            key = product["commodityKey"]
            pname = product["productName"]

            if key.startswith("unknown:"):
                company_data["products"].append({
                    "productName": pname,
                    "commodityKey": key,
                    "latestPrice": None,
                    "status": "unknown"
                })
                continue

            price_data = find_price(key, all_prices)

            if price_data:
                matched += 1
                company_data["products"].append({
                    "productName": pname,
                    "commodityKey": key,
                    "latestPrice": price_data["price"],
                    "unit": price_data.get("unit", ""),
                    "source": price_data.get("source", ""),
                    "status": "matched"
                })
            else:
                company_data["products"].append({
                    "productName": pname,
                    "commodityKey": key,
                    "latestPrice": None,
                    "status": "not_found"
                })
                not_found_list.append(f"{name} ({code}): {pname}")

        results.append(company_data)

    # 6. 输出统计
    rate = matched / total * 100 if total > 0 else 0
    print(f"\n{'=' * 60}")
    print(f"匹配结果:")
    print(f"  公司: {len(companies)}")
    print(f"  产品: {total}")
    print(f"  匹配: {matched}")
    print(f"  未匹配: {total - matched}")
    print(f"  匹配率: {rate:.1f}%")

    if not_found_list:
        print(f"\n未匹配产品 ({len(not_found_list)}):")
        for item in not_found_list[:20]:
            print(f"  - {item}")
        if len(not_found_list) > 20:
            print(f"  ... 还有 {len(not_found_list) - 20} 个")

    # 7. 保存结果
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 公司最新价格
    output = {
        "metadata": {
            "generatedAt": datetime.now().isoformat(),
            "cjSource": os.path.basename(cj_file) if cj_file else None,
            "xyzqSource": os.path.basename(xyzq_file) if xyzq_file else None,
            "totalCompanies": len(companies),
            "totalProducts": total,
            "matchedProducts": matched,
            "matchRate": round(rate, 1)
        },
        "companies": results
    }

    output_file = os.path.join(OUTPUT_DIR, "company-latest-prices.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\n已保存: {output_file}")

    # 价格速查表
    price_table = {}
    for company in results:
        for product in company["products"]:
            if product.get("latestPrice") is not None:
                key = product["commodityKey"]
                if key not in price_table:
                    price_table[key] = {
                        "productName": product["productName"],
                        "price": product["latestPrice"],
                        "unit": product.get("unit", ""),
                        "source": product.get("source", "")
                    }

    table_file = os.path.join(OUTPUT_DIR, "price-table.json")
    with open(table_file, "w", encoding="utf-8") as f:
        json.dump(price_table, f, ensure_ascii=False, indent=2)
    print(f"已保存: {table_file}")

    print(f"\n{'=' * 60}")
    print("更新完成!")


if __name__ == "__main__":
    main()
